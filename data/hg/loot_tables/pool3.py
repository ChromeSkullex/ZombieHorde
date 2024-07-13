import openpyxl
import json
import numpy as np

path = "minecraft_loot.xlsx"

ws = openpyxl.load_workbook(path)['pool3']

ws1 = openpyxl.load_workbook(path)['pool3.1']
ws2 = openpyxl.load_workbook(path)['pool3.2']


def generateWeight(weight):
    match weight:
        case 1:
            return 200
        case 2: 
            return 100
        case 3:
            return 20
        case 4: 
            return 10
        case 5:
            return 5
        
def generateCount(weight):
    match weight:
        case 1:
            return 8
        case 2: 
            return 6
        case 3:
            return 4
        case 4: 
            return 3
        case 5:
            return 3


def genMelee():
    f = open('loot/melee_loot_table.json')
    data = json.load(f)
    data["pools"][0]["entries"] = []

    for row in ws.iter_rows():
        if row[0].value != "melee":
            print(row[0].value, generateWeight(row[1].value))
            data["pools"][0]["entries"].append(
            {
                    "type": "minecraft:item",
                    "name": "apocalypsenow:"+row[0].value,
                    "weight": generateWeight(row[1].value)
            })
        
    with open("loot/melee_loot_table.json", 'w') as wf:
        json.dump(data, wf)
        print(data["pools"][0]["entries"])
genMelee()

def genArmor():
    f = open('loot/armor_loot_table.json')
    data = json.load(f)
    data["pools"][0]["entries"] = []

    for row in ws1.iter_rows():
        if row[0].value != "armor":
            print(row[0].value, generateWeight(row[1].value))
            data["pools"][0]["entries"].append(
            {
                    "type": "minecraft:item",
                    "name": "apocalypsenow:"+row[0].value,
                    "weight": generateWeight(row[1].value)
            })
        
    with open("loot/armor_loot_table.json", 'w') as wf:
        json.dump(data, wf)
        print(data["pools"][0]["entries"])
genArmor()


def genConsume():
    f = open('loot/consumable_loot_table.json')
    data = json.load(f)
    data["pools"][0]["entries"] = []

    for row in ws2.iter_rows():
        if row[0].value != "consumable":
            print(row[0].value, generateWeight(row[1].value))
            data["pools"][0]["entries"].append(
            {
                    "type": "minecraft:item",
                    "name": "apocalypsenow:"+row[0].value,
                    "weight": generateWeight(row[1].value),
                    "functions": [
                        {
                            "function": "minecraft:set_count",
                            "count": {
                                "min": 1,
                                "max": generateCount(row[1].value)
                            }
                        }
                    ]
            })
        
    with open("loot/consumable_loot_table.json", 'w') as wf:
        json.dump(data, wf)
        print(data["pools"][0]["entries"])
genConsume()
