import openpyxl
import json
import numpy as np

path = "minecraft_loot.xlsx"

ws = openpyxl.load_workbook(path)['pool2']

ws1 = openpyxl.load_workbook(path)['pool2.1']
ws2 = openpyxl.load_workbook(path)['pool2.2']
ws3 = openpyxl.load_workbook(path)['pool2.3']

ammos = {}

def generateWeight(weight):
    match weight:
        case 1:
            return 200
        case 2: 
            return 100
        case 3:
            return 20
        case 4: 
            return 10
        case 5:
            return 5
        
def generateCount(weight):
    match weight:
        case 1:
            return 12
        case 2: 
            return 10
        case 3:
            return 7
        case 4: 
            return 4
        case 5:
            return 2


def genAmmo():
    f = open('loot/ammo_loot_table.json')
    data = json.load(f)
    data["pools"][0]["entries"] = []

    for row in ws.iter_rows():
        if row[0].value != "ammo":
            print(row[0].value, generateWeight(row[1].value))
            data["pools"][0]["entries"].append(
            {
                    "type": "minecraft:item",
                    "name": "tac:"+row[0].value,
                    "weight": generateWeight(row[1].value),
                    "functions": [
                        {
                            "function": "minecraft:set_count",
                            "count": {
                                "min": 2,
                                "max": generateCount(row[1].value)
                            }
                        }
                    ]
            })
        
    with open("loot/ammo_loot_table.json", 'w') as wf:
        json.dump(data, wf)
        print(data["pools"][0]["entries"])
genAmmo()

def genGren():
    f = open('loot/grenade_loot_table.json')
    data = json.load(f)
    data["pools"][0]["entries"] = []

    for row in ws1.iter_rows():
        if row[0].value != "grenade":
            print(row[0].value, generateWeight(row[1].value))
            data["pools"][0]["entries"].append(
            {
                    "type": "minecraft:item",
                    "name": "tac:"+row[0].value,
                    "weight": generateWeight(row[1].value),
                    "functions": [
                        {
                            "function": "minecraft:set_count",
                            "count": {
                                "min": 1,
                                "max": 3
                            }
                        }
                    ]
            })
        
    with open("loot/grenade_loot_table.json", 'w') as wf:
        json.dump(data, wf)
        print(data["pools"][0]["entries"])

genGren()
def genScope():
    f = open('loot/scope_loot_table.json')
    data = json.load(f)
    data["pools"][0]["entries"] = []

    for row in ws2.iter_rows():
        if row[0].value != "scope":
            print(row[0].value, generateWeight(row[1].value))
            data["pools"][0]["entries"].append(
            {
                    "type": "minecraft:item",
                    "name": "tac:"+row[0].value,
                    "weight": generateWeight(row[1].value)
            })
        
    with open("loot/scope_loot_table.json", 'w') as wf:
        json.dump(data, wf)
        print(data["pools"][0]["entries"])

genScope()

def genScope():
    f = open('loot/attachment_loot_table.json')
    data = json.load(f)
    data["pools"][0]["entries"] = []

    for row in ws3.iter_rows():
        if row[0].value != "attachment":
            print(row[0].value, generateWeight(row[1].value))
            data["pools"][0]["entries"].append(
            {
                    "type": "minecraft:item",
                    "name": "tac:"+row[0].value,
                    "weight": generateWeight(row[1].value)
            })
        
    with open("loot/attachment_loot_table.json", 'w') as wf:
        json.dump(data, wf)
        print(data["pools"][0]["entries"])

genScope()
